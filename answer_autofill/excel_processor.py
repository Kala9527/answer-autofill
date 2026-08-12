from __future__ import annotations

import asyncio
import shutil
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Protocol

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .config import AppConfig
from .invalid import cell_to_text, is_invalid_answer
from .prompting import QAItem, TargetItem


class GroupCompleter(Protocol):
    async def complete_cell(
        self,
        *,
        basic_info: list[QAItem],
        group_items: list[QAItem],
        context_answers: list[QAItem],
        target: TargetItem,
    ) -> str:
        ...


@dataclass
class SheetStats:
    sheet_name: str
    groups_seen: int = 0
    invalid_cells: int = 0
    filled_cells: int = 0
    failed_cells: int = 0
    warnings: list[str] = field(default_factory=list)


@dataclass
class FileStats:
    source: Path
    output: Path | None
    sheets: list[SheetStats] = field(default_factory=list)

    @property
    def filled_cells(self) -> int:
        return sum(sheet.filled_cells for sheet in self.sheets)

    @property
    def invalid_cells(self) -> int:
        return sum(sheet.invalid_cells for sheet in self.sheets)

    @property
    def failed_cells(self) -> int:
        return sum(sheet.failed_cells for sheet in self.sheets)


@dataclass
class FolderStats:
    input_folder: Path
    output_folder: Path
    files: list[FileStats] = field(default_factory=list)

    @property
    def filled_cells(self) -> int:
        return sum(file.filled_cells for file in self.files)

    @property
    def invalid_cells(self) -> int:
        return sum(file.invalid_cells for file in self.files)

    @property
    def failed_cells(self) -> int:
        return sum(file.failed_cells for file in self.files)


@dataclass
class PendingGroup:
    worksheet: Worksheet
    stats: SheetStats
    basic_info: list[QAItem]
    group_items: list[QAItem]
    valid_answers: list[QAItem]
    targets: list[TargetItem]
    target_columns: dict[str, int]

    @property
    def column_range(self) -> str:
        first = self.group_items[0].column
        last = self.group_items[-1].column
        return f"{first}:{last}"


@dataclass
class SmokeResult:
    source: Path
    sheet_name: str
    group_columns: str
    context_answers: list[QAItem]
    target: TargetItem
    generated: str


class ExcelAnswerProcessor:
    def __init__(self, config: AppConfig, completer: GroupCompleter | None = None, dry_run: bool = False) -> None:
        self.config = config
        self.completer = completer
        self.dry_run = dry_run

    async def process_folder(
        self,
        input_folder: Path,
        output_folder: Path | None = None,
        recursive: bool | None = None,
    ) -> FolderStats:
        input_folder = input_folder.resolve()
        if not input_folder.exists() or not input_folder.is_dir():
            raise FileNotFoundError(f"Input folder does not exist: {input_folder}")

        output_folder = (output_folder or input_folder / self.config.processing.output_dir_name).resolve()
        if not self.dry_run:
            output_folder.mkdir(parents=True, exist_ok=True)

        use_recursive = self.config.processing.recursive if recursive is None else recursive
        stats = FolderStats(input_folder=input_folder, output_folder=output_folder)
        for source in self._iter_excel_files(input_folder, output_folder, use_recursive):
            output = None if self.dry_run else self._next_output_path(source, output_folder)
            stats.files.append(await self.process_file(source, output))
        return stats

    async def process_file(self, source: Path, output: Path | None) -> FileStats:
        source = source.resolve()
        if output is not None:
            output.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(source, output)
            workbook_path = output
        else:
            workbook_path = source

        keep_vba = source.suffix.lower() == ".xlsm"
        workbook = load_workbook(workbook_path, keep_vba=keep_vba, data_only=False)
        file_stats = FileStats(source=source, output=output)
        pending_groups: list[PendingGroup] = []

        for worksheet in workbook.worksheets:
            sheet_stats, sheet_pending = self._scan_sheet(worksheet)
            file_stats.sheets.append(sheet_stats)
            pending_groups.extend(sheet_pending)

        if output is not None:
            await self._complete_pending_groups(pending_groups)
            workbook.save(output)
        return file_stats

    async def smoke_test(self, input_folder: Path, recursive: bool | None = None, smoke_column: str | None = None) -> SmokeResult:
        if self.completer is None:
            raise RuntimeError("Smoke test requires a model completer.")

        input_folder = input_folder.resolve()
        output_folder = input_folder / self.config.processing.output_dir_name
        use_recursive = self.config.processing.recursive if recursive is None else recursive
        for source in self._iter_excel_files(input_folder, output_folder, use_recursive):
            workbook = load_workbook(source, data_only=False)
            for worksheet in workbook.worksheets:
                _, pending_groups = self._scan_sheet(worksheet)
                if not pending_groups:
                    continue
                pending = self._select_smoke_group(pending_groups, smoke_column)
                if pending is None:
                    continue
                target = pending.targets[0]
                if smoke_column:
                    target = next(target for target in pending.targets if target.column == smoke_column.upper())
                context_answers = list(pending.valid_answers)
                generated = await self.completer.complete_cell(
                    basic_info=pending.basic_info,
                    group_items=pending.group_items,
                    context_answers=context_answers,
                    target=target,
                )
                return SmokeResult(
                    source=source,
                    sheet_name=worksheet.title,
                    group_columns=pending.column_range,
                    context_answers=context_answers,
                    target=target,
                    generated=generated,
                )
        raise RuntimeError("No invalid answer group was found for smoke test.")

    @staticmethod
    def _select_smoke_group(pending_groups: list[PendingGroup], smoke_column: str | None) -> PendingGroup | None:
        if not smoke_column:
            return pending_groups[0]
        wanted = smoke_column.upper()
        for pending in pending_groups:
            if any(target.column == wanted for target in pending.targets):
                return pending
        return None

    def _scan_sheet(self, worksheet: Worksheet) -> tuple[SheetStats, list[PendingGroup]]:
        sheet_stats = SheetStats(sheet_name=worksheet.title)
        pending_groups: list[PendingGroup] = []
        excel_cfg = self.config.excel
        header_row = excel_cfg.header_row
        answer_row = excel_cfg.answer_row
        if worksheet.max_row < max(header_row, answer_row):
            return sheet_stats, pending_groups

        basic_info = self._collect_basic_info(worksheet)
        group_start = column_index_from_string(excel_cfg.question_group_start_col)
        group_size = excel_cfg.question_group_size

        for start_col in range(group_start, worksheet.max_column + 1, group_size):
            group_cols = list(range(start_col, min(start_col + group_size, worksheet.max_column + 1)))
            group_items = self._collect_group_items(worksheet, group_cols)
            if not group_items:
                continue
            sheet_stats.groups_seen += 1

            valid_answers: list[QAItem] = []
            targets: list[TargetItem] = []
            target_columns: dict[str, int] = {}
            for item in group_items:
                col_idx = column_index_from_string(item.column)
                original = worksheet.cell(answer_row, col_idx).value
                if is_invalid_answer(original):
                    targets.append(TargetItem(item.column, item.question, cell_to_text(original)))
                    target_columns[item.column] = col_idx
                else:
                    valid_answers.append(item)

            if not targets:
                continue

            sheet_stats.invalid_cells += len(targets)
            if self.dry_run:
                continue
            pending_groups.append(
                PendingGroup(
                    worksheet=worksheet,
                    stats=sheet_stats,
                    basic_info=basic_info,
                    group_items=group_items,
                    valid_answers=valid_answers,
                    targets=targets,
                    target_columns=target_columns,
                )
            )
        return sheet_stats, pending_groups

    async def _complete_pending_groups(self, pending_groups: list[PendingGroup]) -> None:
        if not pending_groups:
            return
        if self.completer is None:
            raise RuntimeError("Non dry-run mode requires a model completer.")

        async def run_one(pending: PendingGroup) -> tuple[PendingGroup, list[tuple[TargetItem, str]] | None, Exception | None]:
            try:
                generated_items: list[tuple[TargetItem, str]] = []
                context_answers = list(pending.valid_answers)
                group_items = list(pending.group_items)
                for target in pending.targets:
                    answer = await self.completer.complete_cell(
                        basic_info=pending.basic_info,
                        group_items=group_items,
                        context_answers=context_answers,
                        target=target,
                    )
                    generated_items.append((target, answer))
                    generated_item = QAItem(target.column, target.question, answer)
                    context_answers.append(generated_item)
                    group_items = [
                        generated_item if item.column == target.column else item
                        for item in group_items
                    ]
                return pending, generated_items, None
            except Exception as exc:  # noqa: BLE001 - keep processing other groups and report exact failure.
                return pending, None, exc

        tasks = [asyncio.create_task(run_one(pending)) for pending in pending_groups]
        for task in asyncio.as_completed(tasks):
            pending, generated, error = await task
            if error is not None:
                pending.stats.failed_cells += len(pending.targets)
                columns = ", ".join(target.column for target in pending.targets)
                pending.stats.warnings.append(f"{pending.worksheet.title} {columns} completion failed: {error}")
                continue

            assert generated is not None
            for target, answer in generated:
                answer = answer.strip()
                if not answer or is_invalid_answer(answer):
                    pending.stats.failed_cells += 1
                    pending.stats.warnings.append(
                        f"{pending.worksheet.title} {target.column} model returned an empty or invalid answer."
                    )
                    continue
                pending.worksheet.cell(self.config.excel.answer_row, pending.target_columns[target.column]).value = answer
                pending.stats.filled_cells += 1

    def _collect_basic_info(self, worksheet: Worksheet) -> list[QAItem]:
        excel_cfg = self.config.excel
        answer_row = excel_cfg.answer_row
        header_row = excel_cfg.header_row
        info: list[QAItem] = []
        for col in excel_cfg.basic_info_cols:
            col_idx = column_index_from_string(col)
            if col_idx > worksheet.max_column:
                continue
            question = cell_to_text(worksheet.cell(header_row, col_idx).value)
            answer = cell_to_text(worksheet.cell(answer_row, col_idx).value)
            if not question or not answer:
                continue
            info.append(QAItem(get_column_letter(col_idx), question, answer))

        occupation_idx = column_index_from_string(excel_cfg.occupation_col)
        occupation_code = cell_to_text(worksheet.cell(answer_row, occupation_idx).value)
        if occupation_code == "1":
            info.append(QAItem(excel_cfg.occupation_col, "occupation meaning", "working"))
        elif occupation_code == "2":
            info.append(QAItem(excel_cfg.occupation_col, "occupation meaning", "student"))
        return info

    def _collect_group_items(self, worksheet: Worksheet, group_cols: list[int]) -> list[QAItem]:
        items: list[QAItem] = []
        for col_idx in group_cols:
            question = cell_to_text(worksheet.cell(self.config.excel.header_row, col_idx).value)
            if not question:
                continue
            answer = cell_to_text(worksheet.cell(self.config.excel.answer_row, col_idx).value)
            items.append(QAItem(get_column_letter(col_idx), question, answer))
        return items

    def _iter_excel_files(self, input_folder: Path, output_folder: Path, recursive: bool) -> list[Path]:
        pattern = "**/*" if recursive else "*"
        files: list[Path] = []
        for path in input_folder.glob(pattern):
            if not path.is_file():
                continue
            if path.name.startswith("~$"):
                continue
            if path.suffix.lower() not in {".xlsx", ".xlsm"}:
                continue
            if self.config.processing.skip_output_folders and self._is_relative_to(path.resolve(), output_folder.resolve()):
                continue
            files.append(path.resolve())
        return sorted(files)

    def _next_output_path(self, source: Path, output_folder: Path) -> Path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        candidate = output_folder / f"{source.stem}_completed_{timestamp}{source.suffix}"
        counter = 1
        while candidate.exists():
            candidate = output_folder / f"{source.stem}_completed_{timestamp}_{counter}{source.suffix}"
            counter += 1
        return candidate

    @staticmethod
    def _is_relative_to(path: Path, parent: Path) -> bool:
        try:
            path.relative_to(parent)
            return True
        except ValueError:
            return False
